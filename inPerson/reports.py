from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

readWorkBook = load_workbook(filename = 'shippingDetails.xlsx')
sheet = readWorkBook['Worksheet'] 

dest_filename = 'emailErrors.xlsx'
wb = Workbook()
ws1 = wb.active

errorCounter = 0

for row in range(2,sheet.max_row+1):  
    for column in "I":
        cell_name = "{}{}".format(column, row)
        email = sheet[cell_name].value
        regResult = re.search("@.+(\.org|\.com|\.edu|\.edu|\.net)$", email)
        if regResult is None:
            print(str(email))
            ws1["A1"] = str(email)
            errorCounter = errorCounter + 1        

wb.save(filename = dest_filename)          